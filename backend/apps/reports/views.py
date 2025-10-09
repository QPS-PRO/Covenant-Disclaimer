from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from rest_framework import status
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Count, Q, Prefetch
from datetime import datetime
import io

# PDF Generation
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    PageBreak,
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT

# Excel Generation
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from apps.assets.models import Employee, Asset, AssetTransaction, Department
from apps.disclaimer.models import DisclaimerProcess, DisclaimerRequest
from apps.disclaimer.permissions import IsAdmin
from rest_framework.permissions import AllowAny

# ============ UTILITY FUNCTIONS ============


def create_pdf_header(elements, title, styles):
    """Create PDF header with title and metadata"""
    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Heading1"],
        fontSize=18,
        textColor=colors.HexColor("#1e40af"),
        spaceAfter=30,
        alignment=TA_CENTER,
    )

    elements.append(Paragraph(title, title_style))
    elements.append(
        Paragraph(
            f"Generated on: {timezone.now().strftime('%B %d, %Y at %H:%M')}",
            ParagraphStyle(
                "Subtitle",
                parent=styles["Normal"],
                alignment=TA_CENTER,
                fontSize=10,
                textColor=colors.grey,
            ),
        )
    )
    elements.append(Spacer(1, 0.3 * inch))


def style_excel_header(ws, headers, row=1):
    """Style Excel header row"""
    header_fill = PatternFill(
        start_color="1e40af", end_color="1e40af", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border


def auto_size_columns(ws):
    """Auto-size Excel columns"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


# ============ DISCLAIMER REPORTS ============


def test_report_simple(request):
    """
    Simple test report without DRF decorators
    """
    from django.http import JsonResponse
    return JsonResponse({"message": "Simple test report working!", "format": request.GET.get("format", "none")})


def disclaimer_completion_report_simple(request):
    """
    Simple disclaimer completion report without DRF decorators
    """
    from django.http import JsonResponse
    format_type = request.GET.get("format", "pdf").lower()
    
    # Get all employees with their disclaimer status
    employees = Employee.objects.select_related("user", "department").all()
    
    completed_count = 0
    not_completed_count = 0
    
    for emp in employees:
        completed_process = emp.disclaimer_processes.filter(status="completed").first()
        if completed_process:
            completed_count += 1
        else:
            not_completed_count += 1
    
    return JsonResponse({
        "message": "Disclaimer completion report (simple)",
        "format": format_type,
        "total_employees": employees.count(),
        "completed": completed_count,
        "not_completed": not_completed_count
    })


def employee_assets_report_simple(request):
    """
    Simple employee assets report without DRF decorators
    """
    from django.http import JsonResponse
    format_type = request.GET.get("format", "pdf").lower()
    
    # Get all employees with their assets
    employees = Employee.objects.select_related("user", "department").prefetch_related("current_assets").all()
    
    with_assets_count = 0
    without_assets_count = 0
    
    for emp in employees:
        current_assets = emp.current_assets.filter(status="assigned")
        if current_assets.exists():
            with_assets_count += 1
        else:
            without_assets_count += 1
    
    return JsonResponse({
        "message": "Employee assets report (simple)",
        "format": format_type,
        "total_employees": employees.count(),
        "with_assets": with_assets_count,
        "without_assets": without_assets_count
    })


def assets_by_status_report_simple(request):
    """
    Simple assets by status report without DRF decorators
    """
    from django.http import JsonResponse
    format_type = request.GET.get("format", "pdf").lower()
    
    # Get all assets grouped by status
    assets = Asset.objects.select_related("department", "current_holder__user").all()
    
    assets_by_status = {
        "available": 0,
        "assigned": 0,
        "maintenance": 0,
        "retired": 0,
    }
    
    for asset in assets:
        assets_by_status[asset.status] += 1
    
    return JsonResponse({
        "message": "Assets by status report (simple)",
        "format": format_type,
        "total_assets": assets.count(),
        "assets_by_status": assets_by_status
    })


def transaction_history_report_simple(request):
    """
    Simple transaction history report without DRF decorators
    """
    from django.http import JsonResponse
    format_type = request.GET.get("format", "pdf").lower()
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")
    
    # Get transactions
    transactions = AssetTransaction.objects.select_related(
        "asset", "employee__user", "employee__department", "processed_by"
    ).all()
    
    if start_date:
        transactions = transactions.filter(transaction_date__gte=start_date)
    if end_date:
        transactions = transactions.filter(transaction_date__lte=end_date)
    
    transactions = transactions.order_by("-transaction_date")
    
    return JsonResponse({
        "message": "Transaction history report (simple)",
        "format": format_type,
        "total_transactions": transactions.count(),
        "start_date": start_date,
        "end_date": end_date
    })


def department_summary_report_simple(request):
    """
    Simple department summary report without DRF decorators
    """
    from django.http import JsonResponse
    format_type = request.GET.get("format", "pdf").lower()
    
    # Get departments
    departments = Department.objects.prefetch_related(
        "employees", "assets", "employees__disclaimer_processes"
    ).all()
    
    dept_summary = []
    for dept in departments:
        total_employees = dept.employees.count()
        completed_disclaimers = (
            dept.employees.filter(disclaimer_processes__status="completed")
            .distinct()
            .count()
        )
        
        dept_summary.append({
            "name": dept.name,
            "manager": dept.manager.get_full_name() if dept.manager else "No Manager",
            "total_employees": total_employees,
            "completed_disclaimers": completed_disclaimers,
            "total_assets": dept.assets.count(),
        })
    
    return JsonResponse({
        "message": "Department summary report (simple)",
        "format": format_type,
        "total_departments": departments.count(),
        "departments": dept_summary
    })


@api_view(["GET"])
def test_report_minimal(request):
    """
    Minimal DRF test report
    """
    return Response({"message": "Minimal DRF test working!", "format": request.GET.get("format", "none")})


@api_view(["GET"])
@permission_classes([])
def test_report(request):
    """
    Simple test report to debug URL routing
    """
    return Response({"message": "Test report working!", "format": request.GET.get("format", "none")})


@api_view(["GET"])
@permission_classes([AllowAny])
def disclaimer_completion_report(request):
    """
    Report showing employees who have completed vs not completed disclaimer process
    Formats: PDF or Excel
    """
    format_type = request.GET.get("format", "pdf").lower()

    # Get all employees with their disclaimer status
    employees = (
        Employee.objects.select_related("user", "department")
        .prefetch_related(
            Prefetch(
                "disclaimer_processes",
                queryset=DisclaimerProcess.objects.filter(status="completed"),
            )
        )
        .all()
    )

    completed_employees = []
    not_completed_employees = []

    for emp in employees:
        emp_data = {
            "employee_id": emp.employee_id,
            "name": emp.name,
            "email": emp.email,
            "department": emp.department.name,
            "phone": emp.phone_number,
        }

        completed_process = emp.disclaimer_processes.filter(status="completed").first()
        if completed_process:
            emp_data["completed_date"] = completed_process.completed_at.strftime(
                "%Y-%m-%d %H:%M"
            )
            emp_data["total_steps"] = completed_process.total_steps
            completed_employees.append(emp_data)
        else:
            active_process = emp.disclaimer_processes.filter(
                status="in_progress"
            ).first()
            if active_process:
                emp_data["status"] = (
                    f"In Progress ({active_process.current_step}/{active_process.total_steps})"
                )
            else:
                emp_data["status"] = "Not Started"
            not_completed_employees.append(emp_data)

    if format_type == "excel":
        return generate_disclaimer_completion_excel(
            completed_employees, not_completed_employees
        )
    else:
        return generate_disclaimer_completion_pdf(
            completed_employees, not_completed_employees
        )


def generate_disclaimer_completion_pdf(completed, not_completed):
    """Generate PDF for disclaimer completion report"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=letter, topMargin=0.5 * inch, bottomMargin=0.5 * inch
    )
    elements = []
    styles = getSampleStyleSheet()

    create_pdf_header(elements, "Disclaimer Completion Report", styles)

    # Summary
    elements.append(
        Paragraph(
            f"<b>Total Employees:</b> {len(completed) + len(not_completed)}",
            styles["Normal"],
        )
    )
    elements.append(
        Paragraph(
            f"<b>Completed:</b> {len(completed)} ({len(completed) / (len(completed) + len(not_completed)) * 100:.1f}%)",
            styles["Normal"],
        )
    )
    elements.append(
        Paragraph(
            f"<b>Not Completed:</b> {len(not_completed)} ({len(not_completed) / (len(completed) + len(not_completed)) * 100:.1f}%)",
            styles["Normal"],
        )
    )
    elements.append(Spacer(1, 0.3 * inch))

    # Completed Employees Table
    if completed:
        elements.append(
            Paragraph("<b>Employees Who Completed Disclaimer</b>", styles["Heading2"])
        )
        elements.append(Spacer(1, 0.1 * inch))

        data = [
            ["Employee ID", "Name", "Department", "Email", "Completed Date", "Steps"]
        ]
        for emp in completed:
            data.append(
                [
                    emp["employee_id"],
                    emp["name"],
                    emp["department"],
                    emp["email"],
                    emp["completed_date"],
                    str(emp["total_steps"]),
                ]
            )

        table = Table(data, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e40af")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("GRID", (0, 0), (-1, -1), 1, colors.grey),
                    ("FONTSIZE", (0, 1), (-1, -1), 8),
                    (
                        "ROWBACKGROUNDS",
                        (0, 1),
                        (-1, -1),
                        [colors.white, colors.HexColor("#f3f4f6")],
                    ),
                ]
            )
        )
        elements.append(table)
        elements.append(PageBreak())

    # Not Completed Employees Table
    if not_completed:
        elements.append(
            Paragraph(
                "<b>Employees Who Haven't Completed Disclaimer</b>", styles["Heading2"]
            )
        )
        elements.append(Spacer(1, 0.1 * inch))

        data = [["Employee ID", "Name", "Department", "Email", "Status"]]
        for emp in not_completed:
            data.append(
                [
                    emp["employee_id"],
                    emp["name"],
                    emp["department"],
                    emp["email"],
                    emp["status"],
                ]
            )

        table = Table(data, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#dc2626")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("GRID", (0, 0), (-1, -1), 1, colors.grey),
                    ("FONTSIZE", (0, 1), (-1, -1), 8),
                    (
                        "ROWBACKGROUNDS",
                        (0, 1),
                        (-1, -1),
                        [colors.white, colors.HexColor("#f3f4f6")],
                    ),
                ]
            )
        )
        elements.append(table)

    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer.read(), content_type="application/pdf")
    response["Content-Disposition"] = (
        f'attachment; filename="disclaimer_completion_report_{timezone.now().strftime("%Y%m%d")}.pdf"'
    )
    return response


def generate_disclaimer_completion_excel(completed, not_completed):
    """Generate Excel for disclaimer completion report"""
    wb = Workbook()

    # Summary Sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary["A1"] = "Disclaimer Completion Report"
    ws_summary["A1"].font = Font(bold=True, size=16)
    ws_summary["A2"] = f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M')}"

    ws_summary["A4"] = "Total Employees:"
    ws_summary["B4"] = len(completed) + len(not_completed)
    ws_summary["A5"] = "Completed:"
    ws_summary["B5"] = len(completed)
    ws_summary["A6"] = "Not Completed:"
    ws_summary["B6"] = len(not_completed)
    ws_summary["A7"] = "Completion Rate:"
    ws_summary["B7"] = (
        f"{len(completed) / (len(completed) + len(not_completed)) * 100:.1f}%"
    )

    # Completed Sheet
    if completed:
        ws_completed = wb.create_sheet("Completed")
        headers = [
            "Employee ID",
            "Name",
            "Department",
            "Email",
            "Phone",
            "Completed Date",
            "Steps",
        ]
        style_excel_header(ws_completed, headers)

        for row_num, emp in enumerate(completed, 2):
            ws_completed.cell(row=row_num, column=1, value=emp["employee_id"])
            ws_completed.cell(row=row_num, column=2, value=emp["name"])
            ws_completed.cell(row=row_num, column=3, value=emp["department"])
            ws_completed.cell(row=row_num, column=4, value=emp["email"])
            ws_completed.cell(row=row_num, column=5, value=emp["phone"])
            ws_completed.cell(row=row_num, column=6, value=emp["completed_date"])
            ws_completed.cell(row=row_num, column=7, value=emp["total_steps"])

        auto_size_columns(ws_completed)

    # Not Completed Sheet
    if not_completed:
        ws_not_completed = wb.create_sheet("Not Completed")
        headers = ["Employee ID", "Name", "Department", "Email", "Phone", "Status"]
        style_excel_header(ws_not_completed, headers)

        for row_num, emp in enumerate(not_completed, 2):
            ws_not_completed.cell(row=row_num, column=1, value=emp["employee_id"])
            ws_not_completed.cell(row=row_num, column=2, value=emp["name"])
            ws_not_completed.cell(row=row_num, column=3, value=emp["department"])
            ws_not_completed.cell(row=row_num, column=4, value=emp["email"])
            ws_not_completed.cell(row=row_num, column=5, value=emp["phone"])
            ws_not_completed.cell(row=row_num, column=6, value=emp["status"])

        auto_size_columns(ws_not_completed)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f'attachment; filename="disclaimer_completion_report_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    )
    return response


# ============ ASSET REPORTS ============


@api_view(["GET"])
@permission_classes([AllowAny])
def assets_by_status_report(request):
    """
    Report showing assets grouped by status
    Formats: PDF or Excel
    """
    format_type = request.GET.get("format", "pdf").lower()

    assets = Asset.objects.select_related("department", "current_holder__user").all()

    # Group by status
    assets_by_status = {
        "available": [],
        "assigned": [],
        "maintenance": [],
        "retired": [],
    }

    for asset in assets:
        asset_data = {
            "name": asset.name,
            "serial_number": asset.serial_number,
            "department": asset.department.name,
            "purchase_date": asset.purchase_date.strftime("%Y-%m-%d")
            if asset.purchase_date
            else "N/A",
            "purchase_cost": f"${asset.purchase_cost}"
            if asset.purchase_cost
            else "N/A",
            "current_holder": asset.current_holder.name
            if asset.current_holder
            else "N/A",
            "description": asset.description[:50] + "..."
            if len(asset.description) > 50
            else asset.description,
        }
        assets_by_status[asset.status].append(asset_data)

    if format_type == "excel":
        return generate_assets_status_excel(assets_by_status)
    else:
        return generate_assets_status_pdf(assets_by_status)


def generate_assets_status_pdf(assets_by_status):
    """Generate PDF for assets by status report"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4, topMargin=0.5 * inch, bottomMargin=0.5 * inch
    )
    elements = []
    styles = getSampleStyleSheet()

    create_pdf_header(elements, "Assets Status Report", styles)

    # Summary
    total_assets = sum(len(assets) for assets in assets_by_status.values())
    elements.append(Paragraph(f"<b>Total Assets:</b> {total_assets}", styles["Normal"]))
    for status_key, assets in assets_by_status.items():
        elements.append(
            Paragraph(
                f"<b>{status_key.title()}:</b> {len(assets)} ({len(assets) / total_assets * 100:.1f}%)",
                styles["Normal"],
            )
        )
    elements.append(Spacer(1, 0.3 * inch))

    # Tables for each status
    status_colors = {
        "available": colors.HexColor("#10b981"),
        "assigned": colors.HexColor("#3b82f6"),
        "maintenance": colors.HexColor("#f59e0b"),
        "retired": colors.HexColor("#6b7280"),
    }

    for status_key, assets in assets_by_status.items():
        if not assets:
            continue

        elements.append(
            Paragraph(f"<b>{status_key.upper()} ASSETS</b>", styles["Heading2"])
        )
        elements.append(Spacer(1, 0.1 * inch))

        data = [
            ["Name", "Serial Number", "Department", "Current Holder", "Purchase Date"]
        ]
        for asset in assets:
            data.append(
                [
                    asset["name"],
                    asset["serial_number"],
                    asset["department"],
                    asset["current_holder"],
                    asset["purchase_date"],
                ]
            )

        table = Table(data, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), status_colors[status_key]),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 9),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("GRID", (0, 0), (-1, -1), 1, colors.grey),
                    ("FONTSIZE", (0, 1), (-1, -1), 7),
                    (
                        "ROWBACKGROUNDS",
                        (0, 1),
                        (-1, -1),
                        [colors.white, colors.HexColor("#f3f4f6")],
                    ),
                ]
            )
        )
        elements.append(table)
        elements.append(Spacer(1, 0.2 * inch))

    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer.read(), content_type="application/pdf")
    response["Content-Disposition"] = (
        f'attachment; filename="assets_status_report_{timezone.now().strftime("%Y%m%d")}.pdf"'
    )
    return response


def generate_assets_status_excel(assets_by_status):
    """Generate Excel for assets by status report"""
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"

    # Summary
    ws_summary["A1"] = "Assets Status Report"
    ws_summary["A1"].font = Font(bold=True, size=16)
    ws_summary["A2"] = f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M')}"

    total_assets = sum(len(assets) for assets in assets_by_status.values())
    ws_summary["A4"] = "Total Assets:"
    ws_summary["B4"] = total_assets

    row = 5
    for status_key, assets in assets_by_status.items():
        ws_summary[f"A{row}"] = f"{status_key.title()}:"
        ws_summary[f"B{row}"] = len(assets)
        row += 1

    # Create sheet for each status
    for status_key, assets in assets_by_status.items():
        if not assets:
            continue

        ws = wb.create_sheet(status_key.title())
        headers = [
            "Name",
            "Serial Number",
            "Department",
            "Current Holder",
            "Purchase Date",
            "Purchase Cost",
            "Description",
        ]
        style_excel_header(ws, headers)

        for row_num, asset in enumerate(assets, 2):
            ws.cell(row=row_num, column=1, value=asset["name"])
            ws.cell(row=row_num, column=2, value=asset["serial_number"])
            ws.cell(row=row_num, column=3, value=asset["department"])
            ws.cell(row=row_num, column=4, value=asset["current_holder"])
            ws.cell(row=row_num, column=5, value=asset["purchase_date"])
            ws.cell(row=row_num, column=6, value=asset["purchase_cost"])
            ws.cell(row=row_num, column=7, value=asset["description"])

        auto_size_columns(ws)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f'attachment; filename="assets_status_report_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    )
    return response


@api_view(["GET"])
@permission_classes([AllowAny])
def employee_assets_report(request):
    """
    Report showing employees with current assets vs no assets
    Formats: PDF or Excel
    """
    format_type = request.GET.get("format", "pdf").lower()

    employees = (
        Employee.objects.select_related("user", "department")
        .prefetch_related("current_assets")
        .all()
    )

    with_assets = []
    without_assets = []

    for emp in employees:
        emp_data = {
            "employee_id": emp.employee_id,
            "name": emp.name,
            "email": emp.email,
            "department": emp.department.name,
            "phone": emp.phone_number,
        }

        current_assets = emp.current_assets.filter(status="assigned")
        if current_assets.exists():
            emp_data["asset_count"] = current_assets.count()
            emp_data["assets"] = ", ".join(
                [f"{a.name} ({a.serial_number})" for a in current_assets]
            )
            with_assets.append(emp_data)
        else:
            without_assets.append(emp_data)

    if format_type == "excel":
        return generate_employee_assets_excel(with_assets, without_assets)
    else:
        return generate_employee_assets_pdf(with_assets, without_assets)


def generate_employee_assets_pdf(with_assets, without_assets):
    """Generate PDF for employee assets report"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=letter, topMargin=0.5 * inch, bottomMargin=0.5 * inch
    )
    elements = []
    styles = getSampleStyleSheet()

    create_pdf_header(elements, "Employee Assets Report", styles)

    # Summary
    total = len(with_assets) + len(without_assets)
    elements.append(Paragraph(f"<b>Total Employees:</b> {total}", styles["Normal"]))
    elements.append(
        Paragraph(
            f"<b>With Assets:</b> {len(with_assets)} ({len(with_assets) / total * 100:.1f}%)",
            styles["Normal"],
        )
    )
    elements.append(
        Paragraph(
            f"<b>Without Assets:</b> {len(without_assets)} ({len(without_assets) / total * 100:.1f}%)",
            styles["Normal"],
        )
    )
    elements.append(Spacer(1, 0.3 * inch))

    # Employees with assets
    if with_assets:
        elements.append(
            Paragraph("<b>Employees With Current Assets</b>", styles["Heading2"])
        )
        elements.append(Spacer(1, 0.1 * inch))

        data = [["Employee ID", "Name", "Department", "Asset Count", "Assets"]]
        for emp in with_assets:
            data.append(
                [
                    emp["employee_id"],
                    emp["name"],
                    emp["department"],
                    str(emp["asset_count"]),
                    emp["assets"],
                ]
            )

        table = Table(
            data,
            repeatRows=1,
            colWidths=[1 * inch, 1.5 * inch, 1.2 * inch, 0.8 * inch, 3 * inch],
        )
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#10b981")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("GRID", (0, 0), (-1, -1), 1, colors.grey),
                    ("FONTSIZE", (0, 1), (-1, -1), 8),
                    (
                        "ROWBACKGROUNDS",
                        (0, 1),
                        (-1, -1),
                        [colors.white, colors.HexColor("#f3f4f6")],
                    ),
                ]
            )
        )
        elements.append(table)
        elements.append(PageBreak())

    # Employees without assets
    if without_assets:
        elements.append(
            Paragraph("<b>Employees Without Current Assets</b>", styles["Heading2"])
        )
        elements.append(Spacer(1, 0.1 * inch))

        data = [["Employee ID", "Name", "Department", "Email", "Phone"]]
        for emp in without_assets:
            data.append(
                [
                    emp["employee_id"],
                    emp["name"],
                    emp["department"],
                    emp["email"],
                    emp["phone"],
                ]
            )

        table = Table(data, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#6b7280")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("GRID", (0, 0), (-1, -1), 1, colors.grey),
                    ("FONTSIZE", (0, 1), (-1, -1), 8),
                    (
                        "ROWBACKGROUNDS",
                        (0, 1),
                        (-1, -1),
                        [colors.white, colors.HexColor("#f3f4f6")],
                    ),
                ]
            )
        )
        elements.append(table)

    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer.read(), content_type="application/pdf")
    response["Content-Disposition"] = (
        f'attachment; filename="employee_assets_report_{timezone.now().strftime("%Y%m%d")}.pdf"'
    )
    return response


def generate_employee_assets_excel(with_assets, without_assets):
    """Generate Excel for employee assets report"""
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"

    # Summary
    ws_summary["A1"] = "Employee Assets Report"
    ws_summary["A1"].font = Font(bold=True, size=16)
    ws_summary["A2"] = f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M')}"

    total = len(with_assets) + len(without_assets)
    ws_summary["A4"] = "Total Employees:"
    ws_summary["B4"] = total
    ws_summary["A5"] = "With Assets:"
    ws_summary["B5"] = len(with_assets)
    ws_summary["A6"] = "Without Assets:"
    ws_summary["B6"] = len(without_assets)

    # With Assets Sheet
    if with_assets:
        ws_with = wb.create_sheet("With Assets")
        headers = [
            "Employee ID",
            "Name",
            "Department",
            "Email",
            "Phone",
            "Asset Count",
            "Assets",
        ]
        style_excel_header(ws_with, headers)

        for row_num, emp in enumerate(with_assets, 2):
            ws_with.cell(row=row_num, column=1, value=emp["employee_id"])
            ws_with.cell(row=row_num, column=2, value=emp["name"])
            ws_with.cell(row=row_num, column=3, value=emp["department"])
            ws_with.cell(row=row_num, column=4, value=emp["email"])
            ws_with.cell(row=row_num, column=5, value=emp["phone"])
            ws_with.cell(row=row_num, column=6, value=emp["asset_count"])
            ws_with.cell(row=row_num, column=7, value=emp["assets"])

        auto_size_columns(ws_with)

    # Without Assets Sheet
    if without_assets:
        ws_without = wb.create_sheet("Without Assets")
        headers = ["Employee ID", "Name", "Department", "Email", "Phone"]
        style_excel_header(ws_without, headers)

        for row_num, emp in enumerate(without_assets, 2):
            ws_without.cell(row=row_num, column=1, value=emp["employee_id"])
            ws_without.cell(row=row_num, column=2, value=emp["name"])
            ws_without.cell(row=row_num, column=3, value=emp["department"])
            ws_without.cell(row=row_num, column=4, value=emp["email"])
            ws_without.cell(row=row_num, column=5, value=emp["phone"])

        auto_size_columns(ws_without)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f'attachment; filename="employee_assets_report_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    )
    return response


# ============ ADDITIONAL RECOMMENDED REPORTS ============


@api_view(["GET"])
@permission_classes([AllowAny])
def asset_transaction_history_report(request):
    """
    Report showing complete asset transaction history
    Includes: issue/return transactions, face verification status
    """
    format_type = request.GET.get("format", "pdf").lower()
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    transactions = AssetTransaction.objects.select_related(
        "asset", "employee__user", "employee__department", "processed_by"
    ).all()

    if start_date:
        transactions = transactions.filter(transaction_date__gte=start_date)
    if end_date:
        transactions = transactions.filter(transaction_date__lte=end_date)

    transactions = transactions.order_by("-transaction_date")

    transaction_data = []
    for txn in transactions:
        transaction_data.append(
            {
                "date": txn.transaction_date.strftime("%Y-%m-%d %H:%M"),
                "type": txn.transaction_type.title(),
                "asset": f"{txn.asset.name} ({txn.asset.serial_number})",
                "employee": txn.employee.name,
                "employee_id": txn.employee.employee_id,
                "department": txn.employee.department.name,
                "processed_by": txn.processed_by.get_full_name()
                if txn.processed_by
                else "System",
                "face_verified": "Yes" if txn.face_verification_success else "No",
                "confidence": f"{txn.face_verification_confidence * 100:.1f}%"
                if txn.face_verification_success
                else "N/A",
                "notes": txn.notes[:50] + "..." if len(txn.notes) > 50 else txn.notes,
            }
        )

    if format_type == "excel":
        return generate_transaction_history_excel(transaction_data)
    else:
        return generate_transaction_history_pdf(transaction_data)


def generate_transaction_history_pdf(transactions):
    """Generate PDF for transaction history"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        topMargin=0.5 * inch,
        bottomMargin=0.5 * inch,
        leftMargin=0.3 * inch,
        rightMargin=0.3 * inch,
    )
    elements = []
    styles = getSampleStyleSheet()

    create_pdf_header(elements, "Asset Transaction History Report", styles)

    elements.append(
        Paragraph(f"<b>Total Transactions:</b> {len(transactions)}", styles["Normal"])
    )
    elements.append(Spacer(1, 0.2 * inch))

    if transactions:
        data = [["Date", "Type", "Asset", "Employee", "Dept", "Verified"]]
        for txn in transactions:
            data.append(
                [
                    txn["date"],
                    txn["type"],
                    txn["asset"],
                    f"{txn['employee']} ({txn['employee_id']})",
                    txn["department"],
                    txn["face_verified"],
                ]
            )

        table = Table(data, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e40af")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 8),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("GRID", (0, 0), (-1, -1), 1, colors.grey),
                    ("FONTSIZE", (0, 1), (-1, -1), 6),
                    (
                        "ROWBACKGROUNDS",
                        (0, 1),
                        (-1, -1),
                        [colors.white, colors.HexColor("#f3f4f6")],
                    ),
                ]
            )
        )
        elements.append(table)

    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer.read(), content_type="application/pdf")
    response["Content-Disposition"] = (
        f'attachment; filename="transaction_history_{timezone.now().strftime("%Y%m%d")}.pdf"'
    )
    return response


def generate_transaction_history_excel(transactions):
    """Generate Excel for transaction history"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    headers = [
        "Date",
        "Type",
        "Asset",
        "Employee",
        "Employee ID",
        "Department",
        "Processed By",
        "Face Verified",
        "Confidence",
        "Notes",
    ]
    style_excel_header(ws, headers)

    for row_num, txn in enumerate(transactions, 2):
        ws.cell(row=row_num, column=1, value=txn["date"])
        ws.cell(row=row_num, column=2, value=txn["type"])
        ws.cell(row=row_num, column=3, value=txn["asset"])
        ws.cell(row=row_num, column=4, value=txn["employee"])
        ws.cell(row=row_num, column=5, value=txn["employee_id"])
        ws.cell(row=row_num, column=6, value=txn["department"])
        ws.cell(row=row_num, column=7, value=txn["processed_by"])
        ws.cell(row=row_num, column=8, value=txn["face_verified"])
        ws.cell(row=row_num, column=9, value=txn["confidence"])
        ws.cell(row=row_num, column=10, value=txn["notes"])

    auto_size_columns(ws)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f'attachment; filename="transaction_history_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    )
    return response


@api_view(["GET"])
@permission_classes([AllowAny])
def department_summary_report(request):
    """
    Comprehensive department summary report
    Shows: employees count, assets count, disclaimer completion rate
    """
    format_type = request.GET.get("format", "pdf").lower()

    departments = Department.objects.prefetch_related(
        "employees", "assets", "employees__disclaimer_processes"
    ).all()

    dept_data = []
    for dept in departments:
        total_employees = dept.employees.count()
        completed_disclaimers = (
            dept.employees.filter(disclaimer_processes__status="completed")
            .distinct()
            .count()
        )

        dept_data.append(
            {
                "name": dept.name,
                "manager": dept.manager.get_full_name()
                if dept.manager
                else "No Manager",
                "total_employees": total_employees,
                "completed_disclaimers": completed_disclaimers,
                "disclaimer_rate": f"{completed_disclaimers / total_employees * 100:.1f}%"
                if total_employees > 0
                else "0%",
                "total_assets": dept.assets.count(),
                "assigned_assets": dept.assets.filter(status="assigned").count(),
                "available_assets": dept.assets.filter(status="available").count(),
                "maintenance_assets": dept.assets.filter(status="maintenance").count(),
            }
        )

    if format_type == "excel":
        return generate_department_summary_excel(dept_data)
    else:
        return generate_department_summary_pdf(dept_data)


def generate_department_summary_pdf(dept_data):
    """Generate PDF for department summary"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=letter, topMargin=0.5 * inch, bottomMargin=0.5 * inch
    )
    elements = []
    styles = getSampleStyleSheet()

    create_pdf_header(elements, "Department Summary Report", styles)

    elements.append(
        Paragraph(f"<b>Total Departments:</b> {len(dept_data)}", styles["Normal"])
    )
    elements.append(Spacer(1, 0.2 * inch))

    data = [
        [
            "Department",
            "Manager",
            "Employees",
            "Disclaimer Rate",
            "Total Assets",
            "Assigned",
            "Available",
        ]
    ]
    for dept in dept_data:
        data.append(
            [
                dept["name"],
                dept["manager"],
                str(dept["total_employees"]),
                dept["disclaimer_rate"],
                str(dept["total_assets"]),
                str(dept["assigned_assets"]),
                str(dept["available_assets"]),
            ]
        )

    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e40af")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 10),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                ("GRID", (0, 0), (-1, -1), 1, colors.grey),
                ("FONTSIZE", (0, 1), (-1, -1), 9),
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -1),
                    [colors.white, colors.HexColor("#f3f4f6")],
                ),
            ]
        )
    )
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer.read(), content_type="application/pdf")
    response["Content-Disposition"] = (
        f'attachment; filename="department_summary_{timezone.now().strftime("%Y%m%d")}.pdf"'
    )
    return response


def generate_department_summary_excel(dept_data):
    """Generate Excel for department summary"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Department Summary"

    headers = [
        "Department",
        "Manager",
        "Total Employees",
        "Completed Disclaimers",
        "Disclaimer Rate",
        "Total Assets",
        "Assigned Assets",
        "Available Assets",
        "Maintenance",
    ]
    style_excel_header(ws, headers)

    for row_num, dept in enumerate(dept_data, 2):
        ws.cell(row=row_num, column=1, value=dept["name"])
        ws.cell(row=row_num, column=2, value=dept["manager"])
        ws.cell(row=row_num, column=3, value=dept["total_employees"])
        ws.cell(row=row_num, column=4, value=dept["completed_disclaimers"])
        ws.cell(row=row_num, column=5, value=dept["disclaimer_rate"])
        ws.cell(row=row_num, column=6, value=dept["total_assets"])
        ws.cell(row=row_num, column=7, value=dept["assigned_assets"])
        ws.cell(row=row_num, column=8, value=dept["available_assets"])
        ws.cell(row=row_num, column=9, value=dept["maintenance_assets"])

    auto_size_columns(ws)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f'attachment; filename="department_summary_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    )
    return response


@api_view(["GET"])
@permission_classes([AllowAny])
def reports_list_view(request):
    """
    GET: List all available reports with descriptions
    """
    reports = [
        {
            "id": "disclaimer-completion",
            "name": "Disclaimer Completion Report",
            "description": "Shows which employees have completed disclaimer process vs those who haven't",
            "endpoint": "/api/reports/disclaimer-completion/",
            "formats": ["pdf", "excel"],
        },
        {
            "id": "employee-assets",
            "name": "Employee Assets Report",
            "description": "Shows employees with current assigned assets vs those without any assets",
            "endpoint": "/api/reports/employee-assets/",
            "formats": ["pdf", "excel"],
        },
        {
            "id": "assets-by-status",
            "name": "Assets by Status Report",
            "description": "Categorizes all assets by their status (available, assigned, maintenance, retired)",
            "endpoint": "/api/reports/assets-by-status/",
            "formats": ["pdf", "excel"],
        },
        {
            "id": "transaction-history",
            "name": "Asset Transaction History",
            "description": "Complete history of all asset transactions with face verification details",
            "endpoint": "/api/reports/transaction-history/",
            "formats": ["pdf", "excel"],
            "parameters": ["start_date (optional)", "end_date (optional)"],
        },
        {
            "id": "department-summary",
            "name": "Department Summary Report",
            "description": "Comprehensive overview of each department including employees, assets, and disclaimer completion",
            "endpoint": "/api/reports/department-summary/",
            "formats": ["pdf", "excel"],
        },
    ]

    return Response(reports)
